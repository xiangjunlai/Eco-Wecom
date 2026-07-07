#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用 xlsx 智能表格生成脚本（配置驱动）

新客户时：修改顶部 CONFIG 块即可，主体代码不用动。
- 客户名、看板标题、服务商
- 6 个 KPI 指标（标签/值/单位/备注）
- 各 Sheet 字段列表
- 状态枚举、示例数据

用法：
  python3 scripts/generate_xlsx.py
  产物在 OUTPUT 路径
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random
from pathlib import Path

# ========== CONFIG：给新客户时改这里 ==========
CONFIG = {
    # 基本信息
    "客户名": "城邦美商",
    "客户名简称": "城邦美商",
    "服务商名": "[服务商名称待填]",
    "看板标题": "城邦美商 · 通用行业打样订单数据看板",
    "版本": "1.0",

    # 6 大 KPI（标签/值/单位/备注）
    "KPI列表": [
        ("订单总数", "128", "单", "本月新增 +23"),
        ("打样中", "34", "款", "本周新增 +8"),
        ("生产中", "47", "单", "其中 3 单超期"),
        ("已发运", "21", "单", "在途 12 单"),
        ("异常订单", "5", "单", "需立即处理"),
        ("本月营收", "¥ 186", "万", "目标 200 万"),
    ],

    # 状态分布
    "状态分布": [
        ("待确认", 8, "8A8A8A"),
        ("打样中", 34, "FF6B35"),
        ("生产中", 47, "1E5AFF"),
        ("已发运", 21, "3D7BFF"),
        ("已完成", 12, "00C853"),
        ("异常", 5, "E53935"),
    ],

    # 紧急待办示例
    "紧急待办": [
        ("PO-2026-038", "Bloomingdale's", "打样中", "2026-05-25", 5, "张三"),
        ("PO-2026-041", "Anthropologie", "生产中", "2026-05-26", 4, "李四"),
        ("PO-2026-045", "Free People", "打样中", "2026-05-28", 2, "王五"),
        ("PO-2026-029", "Urban Outfitters", "生产中", "2026-05-23", 7, "张三"),
        ("PO-2026-052", "Nordstrom", "待确认", "2026-05-30", 0, "李四"),
        ("PO-2026-033", "Anthropologie", "已发运", "2026-05-24", 6, "王五"),
        ("PO-2026-047", "Bloomingdale's", "打样中", "2026-05-29", 1, "李四"),
        ("PO-2026-050", "Free People", "生产中", "2026-05-31", 0, "王五"),
    ],

    # Sheet 2 订单总览配置
    "订单表头": ["订单号", "客户", "PO号", "下单日期", "约定交期", "状态",
                 "总金额", "已收款", "欠款", "跟单员", "备注"],
    "订单列宽": [14, 16, 14, 12, 12, 12, 14, 14, 14, 10, 20],
    "订单状态枚举": "待确认,打样中,生产中,已发运,已完成,已取消,异常",
    "订单客户列表": ["Bloomingdale's", "Anthropologie", "Free People",
                   "Urban Outfitters", "Nordstrom", "J.Crew", "Madewell"],
    "订单跟单员": ["张三", "李四", "王五"],
    "订单条数": 30,
    "订单号前缀": "PO-2026-",

    # Sheet 3 打样进度配置
    "打样表头": ["打样单号", "关联订单", "客户", "款式", "颜色", "尺码",
                 "数量", "打样状态", "工厂", "寄出日期", "客户确认日期", "备注"],
    "打样列宽": [14, 14, 16, 18, 10, 10, 8, 12, 14, 12, 14, 18],
    "打样状态枚举": "待打版,打版中,已寄出,客户确认,客户打回,已确认",
    "款式列表": ["WOMEN BLOUSE", "MEN'S T-SHIRT", "KIDS DRESS", "WOMEN SKIRT", "MEN JACKET"],
    "颜色列表": ["白", "黑", "蓝", "红", "灰"],
    "尺码列表": ["S", "M", "L", "XL"],
    "工厂列表": ["上海一厂", "苏州二厂", "杭州三厂"],
    "打样条数": 20,
    "打样单号前缀": "SM-2026-",

    # Sheet 4 物料追踪配置
    "物料数据": [
        ("M-COT-001", "纯棉面料", "180g/m² 幅宽1.5m", 1200, 500, 800, "PO-2026-038"),
        ("M-SLK-002", "真丝缎面", "22姆米", 300, 400, 0, "PO-2026-041"),
        ("M-LIN-003", "亚麻布", "230g/m²", 800, 300, 200, "PO-2026-045"),
        ("M-WOL-004", "羊毛呢", "380g/m²", 150, 200, 300, "PO-2026-029"),
        ("M-POL-005", "涤纶里布", "210T", 2500, 1000, 0, "PO-2026-052"),
        ("M-BTN-006", "树脂纽扣", "12mm 4眼", 5000, 1000, 0, "PO-2026-038"),
        ("M-ZIP-007", "金属拉链", "5号 银色", 200, 300, 500, "PO-2026-033"),
        ("M-THR-008", "缝纫线", "40/2 白色", 80, 50, 0, "PO-2026-029"),
    ],
    "物料表头": ["物料编码", "名称", "规格", "当前库存", "安全库存", "在途数",
                "预计到货", "关联订单"],
    "物料列宽": [14, 18, 16, 12, 12, 10, 12, 14],

    # Sheet 5 客户档案
    "客户档案": [
        ("Bloomingdale's", "Sarah K.", "sarah@bloomingdales.com", "美东", "女装中高端", 28, 3850000, "A", "VIP 客户"),
        ("Anthropologie", "Emily R.", "emily@anthropologie.com", "美东", "波西米亚风", 22, 2680000, "A", ""),
        ("Free People", "Megan L.", "megan@freepeople.com", "美东", "少女休闲", 18, 1920000, "A", ""),
        ("Urban Outfitters", "Jessica T.", "jessica@urbanoutfitters.com", "美东", "潮流年轻", 15, 1480000, "B", "账期 60 天"),
        ("Nordstrom", "Ashley M.", "ashley@nordstrom.com", "美西", "全品类", 12, 1650000, "A", ""),
        ("J.Crew", "David P.", "david@jcrew.com", "美东", "休闲商务", 8, 820000, "B", "新客户"),
        ("Madewell", "Lauren H.", "lauren@madewell.com", "美东", "牛仔丹宁", 6, 480000, "B", ""),
    ],
    "客户档案表头": ["客户名", "联系人", "联系方式", "区域", "主营品类",
                    "历史订单数", "累计金额", "信用评级", "备注"],
    "客户档案列宽": [18, 10, 16, 12, 14, 12, 14, 10, 18],

    # Sheet 7 月度报表
    "月度数据": [
        ("2026-01", 18, 420000, 25, 2),
        ("2026-02", 15, 380000, 18, 1),
        ("2026-03", 22, 580000, 32, 3),
        ("2026-04", 28, 720000, 38, 4),
        ("2026-05", 35, 1860000, 45, 5),
    ],
}
# ===============================================

# 计算输出路径
BASE = Path("/workspace/定制开发方案生成器")
OUTPUT_DIR = BASE / "examples" / CONFIG["客户名简称"]
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT = OUTPUT_DIR / f'{CONFIG["客户名简称"]}_智能表格交付_V{CONFIG["版本"].split(".")[0]}.xlsx'

# ===== 样式定义 =====
PRIMARY = "1E5AFF"
SUCCESS = "00C853"
WARNING = "FF6B35"
DANGER = "E53935"
BG_SOFT = "F7F8FA"
BORDER = "E5E7EB"

font_header = Font(name="苹方", size=12, bold=True, color="FFFFFF")
font_title = Font(name="苹方", size=14, bold=True, color="1A1A1A")
font_kpi_label = Font(name="苹方", size=11, color="5A5A5A")
font_kpi_value = Font(name="苹方", size=28, bold=True, color=PRIMARY)
font_kpi_unit = Font(name="苹方", size=14, color="5A5A5A")
font_body = Font(name="苹方", size=11, color="1A1A1A")

fill_header = PatternFill("solid", fgColor=PRIMARY)
fill_zebra = PatternFill("solid", fgColor=BG_SOFT)
fill_kpi_bg = PatternFill("solid", fgColor="FFFFFF")
fill_success = PatternFill("solid", fgColor="E5F7EC")
fill_warning = PatternFill("solid", fgColor="FFF4E5")
fill_danger = PatternFill("solid", fgColor="FFE5E5")

thin = Side(border_style="thin", color=BORDER)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

wb = Workbook()


# ============================================================
# Sheet 1: 数据看板
# ============================================================
ws1 = wb.active
ws1.title = "📊 数据看板"
ws1.sheet_view.showGridLines = False

for col, w in zip("ABCDEFGHIJ", [4, 22, 18, 18, 18, 18, 18, 4, 4, 4]):
    ws1.column_dimensions[col].width = w

ws1["B2"] = CONFIG["看板标题"]
ws1["B2"].font = Font(name="苹方", size=20, bold=True, color="1A1A1A")
ws1.merge_cells("B2:G2")
ws1.row_dimensions[2].height = 36

ws1["B3"] = f"数据更新于 {datetime.now().strftime('%Y-%m-%d %H:%M')} · {CONFIG['服务商名']}出品"
ws1["B3"].font = Font(name="苹方", size=10, color="8A8A8A")
ws1.merge_cells("B3:G3")

# KPI 区域
for i, (label, val, unit, note) in enumerate(CONFIG["KPI列表"]):
    col = 2 + (i % 3) * 2
    row = 5 + (i // 3) * 8
    ws1.cell(row=row, column=col, value=label).font = font_kpi_label
    ws1.cell(row=row+1, column=col, value=val).font = font_kpi_value
    ws1.cell(row=row+1, column=col+1, value=unit).font = font_kpi_unit
    ws1.cell(row=row+2, column=col, value=note).font = Font(name="苹方", size=10, color="8A8A8A")
    ws1.merge_cells(start_row=row, end_row=row, start_column=col, end_column=col+1)
    ws1.merge_cells(start_row=row+1, end_row=row+1, start_column=col, end_column=col+1)
    ws1.merge_cells(start_row=row+2, end_row=row+2, start_column=col, end_column=col+1)
    for r in range(row, row+3):
        ws1.row_dimensions[r].height = 24 if r != row+1 else 36

# 待办
ws1["B18"] = "⚠️ 紧急待办"
ws1["B18"].font = Font(name="苹方", size=14, bold=True, color="1A1A1A")
ws1.merge_cells("B18:G18")

todo_headers = ["订单号", "客户", "状态", "约定交期", "超期天数", "跟单员"]
for j, h in enumerate(todo_headers):
    c = ws1.cell(row=19, column=2+j, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all

todos = CONFIG["紧急待办"]
for i, t in enumerate(todos):
    r = 20 + i
    for j, v in enumerate(t):
        c = ws1.cell(row=r, column=2+j, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
    if t[4] > 3:
        ws1.cell(row=r, column=6).fill = fill_danger
        ws1.cell(row=r, column=6).font = Font(name="苹方", size=11, bold=True, color=DANGER)

# 状态分布
ws1["I18"] = "状态分布"
ws1["I18"].font = Font(name="苹方", size=14, bold=True, color="1A1A1A")
ws1["I18"].alignment = center
ws1.merge_cells("I18:J18")

for i, (s, n, color) in enumerate(CONFIG["状态分布"]):
    r = 19 + i
    c1 = ws1.cell(row=r, column=9, value=s)
    c1.font = Font(name="苹方", size=11, color=color, bold=True)
    c1.alignment = center
    c2 = ws1.cell(row=r, column=10, value=n)
    c2.font = Font(name="苹方", size=11, bold=True)
    c2.alignment = center
    if i % 2 == 1:
        c1.fill = fill_zebra
        c2.fill = fill_zebra

ws1.freeze_panes = "A4"

# ============================================================
# Sheet 2: 订单总览
# ============================================================
ws2 = wb.create_sheet("📦 订单总览")
ws2.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["订单表头"], CONFIG["订单列宽"])):
    c = ws2.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws2.column_dimensions[get_column_letter(j+1)].width = w
ws2.row_dimensions[1].height = 32

status_enum = CONFIG["订单状态枚举"].split(",")
customers = CONFIG["订单客户列表"]
followers = CONFIG["订单跟单员"]

random.seed(42)
n = CONFIG["订单条数"]
for i in range(n):
    r = i + 2
    order_no = f"{CONFIG['订单号前缀']}{i+1:03d}"
    customer = random.choice(customers)
    po = f"CUS-{random.randint(10000, 99999)}"
    order_date = datetime(2026, 4, 1) + timedelta(days=random.randint(0, 50))
    delivery = order_date + timedelta(days=random.randint(30, 60))
    status = random.choice(status_enum)
    total = random.randint(30000, 300000)
    paid = total if status == "已完成" else (total * random.choice([0, 0.3, 0.5, 0.7, 1.0]) if status in ["已发运", "生产中"] else 0)
    paid = round(paid, 2)
    follower = random.choice(followers)
    remark = "" if i % 3 != 0 else "客户要求加急"

    values = [order_no, customer, po, order_date.strftime("%Y-%m-%d"),
              delivery.strftime("%Y-%m-%d"), status, total, paid,
              f"=G{r}-H{r}", follower, remark]
    for j, v in enumerate(values):
        c = ws2.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j not in [10] else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j in [6, 7, 8]:
            c.number_format = '¥ #,##0.00'
            c.alignment = right_align
        if j in [3, 4]:
            c.number_format = 'yyyy-mm-dd'

dv = DataValidation(type="list", formula1=f'"{CONFIG["订单状态枚举"]}"', allow_blank=True)
dv.add(f"F2:F{n+1}")
ws2.add_data_validation(dv)

# 条件格式
red_fill = PatternFill("solid", fgColor="FFE5E5")
red_font = Font(name="苹方", size=11, color=DANGER, bold=True)
ws2.conditional_formatting.add(
    f"E2:E{n+1}",
    FormulaRule(formula=[f'$E2<DATE(2026,6,1)'], fill=red_fill, font=red_font)
)
ws2.conditional_formatting.add(
    f"F2:F{n+1}",
    CellIsRule(operator="equal", formula=['"已完成"'], fill=fill_success,
               font=Font(name="苹方", size=11, color=SUCCESS, bold=True))
)
ws2.conditional_formatting.add(
    f"F2:F{n+1}",
    CellIsRule(operator="equal", formula=['"异常"'], fill=fill_danger,
               font=Font(name="苹方", size=11, color=DANGER, bold=True))
)
ws2.conditional_formatting.add(
    f"F2:F{n+1}",
    CellIsRule(operator="equal", formula=['"打样中"'], fill=fill_warning,
               font=Font(name="苹方", size=11, color=WARNING, bold=True))
)

ws2.freeze_panes = "B2"
ws2.auto_filter.ref = f"A1:K{n+1}"

# ============================================================
# Sheet 3: 打样进度
# ============================================================
ws3 = wb.create_sheet("🧪 打样进度")
ws3.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["打样表头"], CONFIG["打样列宽"])):
    c = ws3.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws3.column_dimensions[get_column_letter(j+1)].width = w
ws3.row_dimensions[1].height = 32

sample_status = CONFIG["打样状态枚举"].split(",")
for i in range(CONFIG["打样条数"]):
    r = i + 2
    sample_no = f"{CONFIG['打样单号前缀']}{i+1:03d}"
    order_ref = f"{CONFIG['订单号前缀']}{random.randint(1, n):03d}"
    cust = random.choice(customers)
    style = random.choice(CONFIG["款式列表"])
    color = random.choice(CONFIG["颜色列表"])
    size = "/".join(random.sample(CONFIG["尺码列表"], random.randint(1, 3)))
    qty = random.randint(20, 200)
    status = random.choice(sample_status)
    factory = random.choice(CONFIG["工厂列表"])
    sent = (datetime(2026, 4, 1) + timedelta(days=random.randint(0, 50))).strftime("%Y-%m-%d")
    confirmed = "" if status in ["待打版", "打版中", "已寄出"] else \
                (datetime(2026, 4, 1) + timedelta(days=random.randint(10, 60))).strftime("%Y-%m-%d")
    remark = "" if i % 4 != 0 else random.choice(["面料调整", "尺寸偏小", "颜色待定", "客户满意"])

    values = [sample_no, order_ref, cust, style, color, size, qty, status,
              factory, sent, confirmed, remark]
    for j, v in enumerate(values):
        c = ws3.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j in [9, 10] and v:
            c.number_format = 'yyyy-mm-dd'

ws3.conditional_formatting.add(
    f"H2:H{CONFIG['打样条数']+1}",
    CellIsRule(operator="equal", formula=['"已确认"'], fill=fill_success,
               font=Font(name="苹方", size=11, color=SUCCESS, bold=True))
)
ws3.conditional_formatting.add(
    f"H2:H{CONFIG['打样条数']+1}",
    CellIsRule(operator="equal", formula=['"客户打回"'], fill=fill_danger,
               font=Font(name="苹方", size=11, color=DANGER, bold=True))
)
ws3.freeze_panes = "B2"
ws3.auto_filter.ref = f"A1:L{CONFIG['打样条数']+1}"

# ============================================================
# Sheet 4: 物料追踪
# ============================================================
ws4 = wb.create_sheet("🏭 物料追踪")
ws4.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["物料表头"], CONFIG["物料列宽"])):
    c = ws4.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws4.column_dimensions[get_column_letter(j+1)].width = w
ws4.row_dimensions[1].height = 32

materials = CONFIG["物料数据"]
for i, m in enumerate(materials):
    r = i + 2
    arrival = (datetime(2026, 6, 1) + timedelta(days=random.randint(0, 30))).strftime("%Y-%m-%d") if m[5] > 0 else ""
    # m = (编码, 名称, 规格, 当前库存, 安全库存, 在途, 关联订单)
    values = [m[0], m[1], m[2], m[3], m[4], m[5], arrival, m[6]]
    for j, v in enumerate(values):
        c = ws4.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j in [6] and v:
            c.number_format = 'yyyy-mm-dd'

ws4.conditional_formatting.add(
    f"D2:E{len(materials)+1}",
    FormulaRule(formula=[f'$D2<$E2'], fill=fill_danger,
                font=Font(name="苹方", size=11, color=DANGER, bold=True))
)
ws4.freeze_panes = "B2"

# ============================================================
# Sheet 5: 客户档案
# ============================================================
ws5 = wb.create_sheet("👥 客户档案")
ws5.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["客户档案表头"], CONFIG["客户档案列宽"])):
    c = ws5.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws5.column_dimensions[get_column_letter(j+1)].width = w
ws5.row_dimensions[1].height = 32

for i, c_data in enumerate(CONFIG["客户档案"]):
    r = i + 2
    for j, v in enumerate(c_data):
        c = ws5.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j not in [8] else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j == 6:
            c.number_format = '¥ #,##0'

ws5.conditional_formatting.add(
    f"H2:H{len(CONFIG['客户档案'])+1}",
    CellIsRule(operator="equal", formula=['"A"'], fill=fill_success,
               font=Font(name="苹方", size=11, color=SUCCESS, bold=True))
)
ws5.conditional_formatting.add(
    f"H2:H{len(CONFIG['客户档案'])+1}",
    CellIsRule(operator="equal", formula=['"B"'], fill=fill_warning,
               font=Font(name="苹方", size=11, color=WARNING, bold=True))
)
ws5.freeze_panes = "B2"

# ============================================================
# Sheet 6: 财务对账
# ============================================================
ws6 = wb.create_sheet("💰 财务对账")
ws6.sheet_view.showGridLines = False

fin_headers = ["订单号", "客户", "应收金额", "已收", "未收", "收款日期", "发票号", "备注"]
widths6 = [14, 16, 14, 14, 14, 12, 14, 18]
for j, (h, w) in enumerate(zip(fin_headers, widths6)):
    c = ws6.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws6.column_dimensions[get_column_letter(j+1)].width = w
ws6.row_dimensions[1].height = 32

for i in range(15):
    r = i + 2
    order_no = f"{CONFIG['订单号前缀']}{i+1:03d}"
    cust = random.choice(customers)
    total = random.randint(50000, 300000)
    paid_choices = [0, total*0.3, total*0.5, total*0.7, total]
    paid = round(random.choice(paid_choices), 2)
    receive_date = "" if paid == 0 else (datetime(2026, 4, 1) + timedelta(days=random.randint(0, 50))).strftime("%Y-%m-%d")
    invoice = "" if paid == 0 else f"INV-2026-{random.randint(1000, 9999)}"
    remark = "全款已结" if paid == total else ("部分收款" if paid > 0 else "待收款")

    values = [order_no, cust, total, paid, f"=C{r}-D{r}", receive_date, invoice, remark]
    for j, v in enumerate(values):
        c = ws6.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j in [2, 3, 4]:
            c.number_format = '¥ #,##0.00'
            c.alignment = right_align
        if j == 5 and v:
            c.number_format = 'yyyy-mm-dd'

ws6.conditional_formatting.add(
    f"E2:E{16}",
    FormulaRule(formula=[f'$E2>0'], fill=fill_warning,
                font=Font(name="苹方", size=11, color=WARNING, bold=True))
)

# 汇总行
r = 17
ws6.cell(row=r, column=2, value="合计")
ws6.cell(row=r, column=2).fill = fill_header
ws6.cell(row=r, column=2).font = font_header
ws6.cell(row=r, column=2).alignment = center
for j, col in enumerate([3, 4, 5]):
    c = ws6.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}16)")
    c.number_format = '¥ #,##0.00'
    c.alignment = right_align
    c.font = Font(name="苹方", size=12, bold=True)
    c.fill = fill_warning
for col in [6, 7, 8]:
    ws6.cell(row=r, column=col).fill = fill_header

ws6.freeze_panes = "B2"

# ============================================================
# Sheet 7: 月度报表
# ============================================================
ws7 = wb.create_sheet("📈 月度报表")
ws7.sheet_view.showGridLines = False

ws7["B2"] = f"{datetime.now().year} 年月度业绩报表"
ws7["B2"].font = Font(name="苹方", size=16, bold=True)
ws7.merge_cells("B2:F2")

monthly_headers = ["月份", "订单数", "营收", "打样数", "超期数", "环比"]
for j, h in enumerate(monthly_headers):
    c = ws7.cell(row=4, column=2+j, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws7.column_dimensions[get_column_letter(j+2)].width = 14
ws7.row_dimensions[4].height = 32

monthly = CONFIG["月度数据"]
for i, m in enumerate(monthly):
    r = 5 + i
    ws7.cell(row=r, column=2, value=m[0])
    ws7.cell(row=r, column=3, value=m[1])
    ws7.cell(row=r, column=4, value=m[2]).number_format = '¥ #,##0'
    ws7.cell(row=r, column=5, value=m[3])
    ws7.cell(row=r, column=6, value=m[4])
    if i > 0:
        ws7.cell(row=r, column=7, value=f"=(D{r}-D{r-1})/D{r-1}").number_format = '0.0%'
    for col in range(2, 8):
        c = ws7.cell(row=r, column=col)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra

wb.save(OUTPUT)
print(f"✅ 已生成: {OUTPUT}")
print(f"   包含 7 个 Sheet: 数据看板 / 订单总览 / 打样进度 / 物料追踪 / 客户档案 / 财务对账 / 月度报表")
