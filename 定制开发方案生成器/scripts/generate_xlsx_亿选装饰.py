#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
亿选装饰 xlsx 智能表格生成脚本（家装 / 前端数据分析场景）

由通用 generate_xlsx.py 复制而来，CONFIG 全部替换为家装场景：
- Sheet 1 数据看板：6 大 KPI + 紧急待办 + 状态分布
- Sheet 2 数据源管理：连接 8 个数据源（线上/线下/广告/工长群/材料商/财务/老客/复购）
- Sheet 3 维度配置：12 个分析维度（来源/区域/户型/面积/客单价/工期/复购率/流失率等）
- Sheet 4 工地管理：47 个在施工地的实时进度跟踪
- Sheet 5 客户档案：贵州本地 7 类业主画像
- Sheet 6 异常预警：本周需处理的 12 个异常工单
- Sheet 7 趋势分析：近 6 月前端数据走势 + 环比

新客户时参考顶部 CONFIG 块即可。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random
from pathlib import Path

# ========== CONFIG：给新客户时改这里 ==========
CONFIG = {
    # 基本信息
    "客户名": "亿选装饰",
    "客户名简称": "亿选装饰",
    "服务商名": "杭州装企云网络科技有限公司",
    "看板标题": "亿选装饰 · 家装前端数据驾驶舱",
    "版本": "1.0",
    "场景": "家装前端数据分析",

    # 6 大 KPI（标签/值/单位/备注）
    "KPI列表": [
        ("在施工地", "127", "个", "本周新增 23 个"),
        ("待签约", "34", "户", "本周新增 8 户"),
        ("本月签约", "47", "户", "环比 +12%"),
        ("线索来源", "21", "条", "线上 12 / 线下 9"),
        ("异常工地", "5", "个", "需立即处理"),
        ("本月营收", "¥ 186", "万", "目标 220 万"),
    ],

    # 状态分布
    "状态分布": [
        ("咨询中", 28, "8A8A8A"),
        ("量房中", 34, "FF6B35"),
        ("报价中", 47, "1E5AFF"),
        ("施工中", 121, "3D7BFF"),
        ("已竣工", 12, "00C853"),
        ("已流失", 5, "E53935"),
    ],

    # 紧急待办（工地预警）
    "紧急待办": [
        ("GZ-2026-038", "王女士 / 云岩区", "施工中-延期", "2026-05-25", 5, "张工长"),
        ("GZ-2026-041", "李先生 / 观山湖", "施工中-延期", "2026-05-26", 4, "李工长"),
        ("GZ-2026-045", "张总 / 遵义", "泥木-验收", "2026-05-28", 2, "王工长"),
        ("GZ-2026-029", "陈女士 / 南明区", "油漆-超期", "2026-05-23", 7, "张工长"),
        ("GZ-2026-052", "刘先生 / 六盘水", "待签约", "2026-05-30", 0, "李工长"),
        ("GZ-2026-033", "赵女士 / 花溪区", "软装-急件", "2026-05-24", 6, "王工长"),
        ("GZ-2026-047", "孙先生 / 毕节", "隐蔽-验收", "2026-05-29", 1, "李工长"),
        ("GZ-2026-050", "王女士 / 复购", "已签约-待开工", "2026-05-31", 0, "王工长"),
    ],

    # Sheet 2 数据源管理配置
    "数据源表头": ["数据源", "类型", "接入方式", "更新频率", "字段数", "负责人", "状态", "备注"],
    "数据源列宽": [22, 14, 18, 12, 10, 10, 12, 22],
    "数据源列表": [
        ("巨量引擎投放后台", "广告投放", "API 自动同步", "每小时", 18, "张投放", "✅ 正常", "覆盖抖音/今日头条"),
        ("百度营销后台", "广告投放", "API 自动同步", "每小时", 14, "张投放", "✅ 正常", "覆盖百度搜索/品专"),
        ("美团/大众点评", "本地生活", "API 自动同步", "每日", 9, "陈运营", "✅ 正常", "本地客户主要来源"),
        ("微信客服会话", "在线咨询", "Webhook 接入", "实时", 11, "李客服", "✅ 正常", "需绑定企业微信"),
        ("小程序留资", "在线表单", "Webhook 接入", "实时", 8, "陈运营", "✅ 正常", "官网/小程序入口"),
        ("线下到店登记", "门店", "手工录入", "每日", 12, "前台", "⚠️ 滞后", "建议改为扫码登记"),
        ("老客户转介绍", "口碑", "手工录入", "每周", 6, "全员", "✅ 正常", "复购主要来源"),
        ("材料商返单", "B 端", "API 自动同步", "每日", 5, "采购部", "✅ 正常", "关联工地数据"),
        ("财务收款系统", "财务", "API 自动同步", "实时", 7, "财务部", "✅ 正常", "金蝶云"),
        ("工长群日报", "施工", "模板消息", "每日", 15, "工程部", "✅ 正常", "工长拍照打卡"),
    ],
    "数据源类型枚举": "广告投放,本地生活,在线咨询,在线表单,门店,口碑,B 端,财务,施工",
    "数据源状态枚举": "✅ 正常,⚠️ 滞后,❌ 异常,🔧 维护中",
    "数据源条数": 10,

    # Sheet 3 维度配置
    "维度表头": ["维度名称", "维度类型", "枚举值/口径", "用途", "默认显示", "维护人"],
    "维度列宽": [20, 14, 36, 28, 12, 10],
    "维度列表": [
        ("客户来源", "枚举", "线上广告 / 线下到店 / 转介绍 / 美团 / 自然到店", "投放 ROI 分析", "✅", "运营"),
        ("客户区域", "枚举", "贵阳 / 遵义 / 六盘水 / 毕节 / 安顺 / 黔南", "区域市场分布", "✅", "运营"),
        ("户型", "枚举", "两室 / 三室 / 四室 / 跃层 / 别墅 / 复式", "户型偏好分析", "✅", "设计"),
        ("面积段", "区间", "60-90 / 90-120 / 120-150 / 150-200 / 200+", "客单价分层", "✅", "设计"),
        ("装修风格", "枚举", "现代 / 北欧 / 新中式 / 轻奢 / 极简 / 美式", "风格偏好", "⬜", "设计"),
        ("预算段", "区间", "10 万内 / 10-20 万 / 20-30 万 / 30-50 万 / 50 万+", "客单价分层", "✅", "销售"),
        ("工期", "区间", "60 天内 / 60-90 天 / 90-120 天 / 120+ 天", "工期效率", "✅", "工程"),
        ("阶段", "枚举", "咨询 / 量房 / 报价 / 签约 / 施工 / 竣工", "漏斗分析", "✅", "销售"),
        ("客户等级", "枚举", "A / B / C / D", "客户分级", "✅", "销售"),
        ("签约次数", "区间", "首次 / 复购 1 次 / 复购 2+", "复购分析", "⬜", "运营"),
        ("转介绍路径", "枚举", "老客户 / 设计师 / 工长 / 材料商 / 异业", "口碑分析", "⬜", "运营"),
        ("流失原因", "枚举", "价格 / 工期 / 竞品 / 决策延期 / 其他", "流失分析", "⬜", "销售"),
    ],
    "维度条数": 12,

    # Sheet 4 工地管理
    "工地表头": ["工地编号", "客户", "区域", "户型", "面积", "阶段", "工期进度",
                 "工长", "设计师", "开工日", "预计竣工", "状态", "异常标记"],
    "工地列宽": [14, 12, 12, 12, 8, 14, 12, 8, 10, 12, 12, 12, 14],
    "工地阶段枚举": "拆除,水电,泥木,油漆,软装,已竣工,待开工",
    "工地状态枚举": "✅ 正常,⚠️ 延期,❌ 停工,🔧 整改",
    "工地区域列表": ["云岩区", "南明区", "观山湖区", "花溪区", "遵义", "六盘水", "毕节", "安顺"],
    "工地户型列表": ["两室", "三室", "四室", "跃层", "别墅", "复式"],
    "工长列表": ["张工长", "李工长", "王工长", "赵工长"],
    "设计师列表": ["陈设计", "刘设计", "黄设计", "周设计"],
    "工地条数": 30,
    "工地号前缀": "GZ-2026-",

    # Sheet 5 客户档案
    "客户档案": [
        ("王女士", "13900000001", "13900000001", "贵阳云岩区", "140m²三室两厅", 1, 280000, "A", "高端装修"),
        ("李先生", "13900000002", "13900000002", "贵阳观山湖区", "180m²跃层", 1, 480000, "A", "转介绍客户"),
        ("张总", "13900000003", "13900000003", "遵义汇川区", "260m²别墅", 2, 850000, "A", "老客户复购"),
        ("陈女士", "13900000004", "13900000004", "贵阳南明区", "95m²两室", 1, 165000, "B", "线上投放"),
        ("刘先生", "13900000005", "13900000005", "六盘水钟山区", "120m²三室", 1, 220000, "A", "自然到店"),
        ("赵女士", "13900000006", "13900000006", "贵阳花溪区", "200m²四室", 1, 380000, "B", "美团引流"),
        ("孙先生", "13900000007", "13900000007", "毕节七星关区", "150m²三室", 1, 295000, "A", "老客户介绍"),
        ("周总", "13900000008", "13900000008", "安顺西秀区", "320m²别墅", 1, 1200000, "A", "高端定制"),
    ],
    "客户档案表头": ["客户名", "电话", "联系方式", "区域", "户型",
                    "签约次数", "累计金额", "客户等级", "客户来源"],
    "客户档案列宽": [12, 14, 14, 14, 14, 10, 14, 10, 14],

    # Sheet 6 异常预警
    "异常预警": [
        ("GZ-2026-029", "陈女士 / 南明", "油漆超期", "7 天", "张工长", "需追加工人"),
        ("GZ-2026-038", "王女士 / 云岩", "材料未到", "3 天", "张工长", "瓷砖延期"),
        ("GZ-2026-041", "李先生 / 观山湖", "业主投诉", "2 天", "李工长", "墙面返工"),
        ("GZ-2026-033", "赵女士 / 花溪", "软装延期", "5 天", "王工长", "家具未到货"),
        ("GZ-2026-045", "张总 / 遵义", "泥木验收", "1 天", "王工长", "待业主确认"),
        ("GZ-2026-052", "刘先生 / 六盘水", "合同待签", "0 天", "李工长", "客户出差"),
    ],
    "异常预警表头": ["工地编号", "客户/区域", "异常类型", "已超期", "负责人", "处理建议"],
    "异常预警列宽": [14, 18, 16, 10, 10, 22],

    # Sheet 7 趋势分析
    "月度数据": [
        ("2026-01", 12, 280000, 25, 2, 8.3),
        ("2026-02", 15, 320000, 28, 1, 14.3),
        ("2026-03", 18, 420000, 32, 3, 31.3),
        ("2026-04", 22, 580000, 38, 4, 38.1),
        ("2026-05", 28, 1860000, 45, 5, 220.7),
    ],
}
# ===============================================

# 计算输出路径
BASE = Path("/workspace/定制开发方案生成器")
OUTPUT_DIR = BASE / "examples" / CONFIG["客户名简称"]
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT = OUTPUT_DIR / f'{CONFIG["客户名简称"]}_智能表格交付_V{CONFIG["版本"].split(".")[0]}.xlsx'

# ===== 样式定义 =====
PRIMARY = "1E5AFF"
SUCCESS = "00C853"
WARNING = "FF6B35"
DANGER = "E53935"
BG_SOFT = "F7F8FA"
BORDER = "E5E7EB"

font_header = Font(name="苹方", size=12, bold=True, color="FFFFFF")
font_title = Font(name="苹方", size=14, bold=True, color="1A1A1A")
font_kpi_label = Font(name="苹方", size=11, color="5A5A5A")
font_kpi_value = Font(name="苹方", size=28, bold=True, color=PRIMARY)
font_kpi_unit = Font(name="苹方", size=14, color="5A5A5A")
font_body = Font(name="苹方", size=11, color="1A1A1A")

fill_header = PatternFill("solid", fgColor=PRIMARY)
fill_zebra = PatternFill("solid", fgColor=BG_SOFT)
fill_kpi_bg = PatternFill("solid", fgColor="FFFFFF")
fill_success = PatternFill("solid", fgColor="E5F7EC")
fill_warning = PatternFill("solid", fgColor="FFF4E5")
fill_danger = PatternFill("solid", fgColor="FFE5E5")

thin = Side(border_style="thin", color=BORDER)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

wb = Workbook()


# ============================================================
# Sheet 1: 数据看板
# ============================================================
ws1 = wb.active
ws1.title = "📊 数据看板"
ws1.sheet_view.showGridLines = False

for col, w in zip("ABCDEFGHIJ", [4, 22, 18, 18, 18, 18, 18, 4, 14, 14]):
    ws1.column_dimensions[col].width = w

ws1["B2"] = CONFIG["看板标题"]
ws1["B2"].font = Font(name="苹方", size=20, bold=True, color="1A1A1A")
ws1.merge_cells("B2:G2")
ws1.row_dimensions[2].height = 36

ws1["B3"] = f"数据更新于 {datetime.now().strftime('%Y-%m-%d %H:%M')} · {CONFIG['服务商名']}出品"
ws1["B3"].font = Font(name="苹方", size=10, color="8A8A8A")
ws1.merge_cells("B3:G3")

# KPI 区域
for i, (label, val, unit, note) in enumerate(CONFIG["KPI列表"]):
    col = 2 + (i % 3) * 2
    row = 5 + (i // 3) * 8
    ws1.cell(row=row, column=col, value=label).font = font_kpi_label
    ws1.cell(row=row+1, column=col, value=val).font = font_kpi_value
    ws1.cell(row=row+1, column=col+1, value=unit).font = font_kpi_unit
    ws1.cell(row=row+2, column=col, value=note).font = Font(name="苹方", size=10, color="8A8A8A")
    ws1.merge_cells(start_row=row, end_row=row, start_column=col, end_column=col+1)
    ws1.merge_cells(start_row=row+1, end_row=row+1, start_column=col, end_column=col+1)
    ws1.merge_cells(start_row=row+2, end_row=row+2, start_column=col, end_column=col+1)
    for r in range(row, row+3):
        ws1.row_dimensions[r].height = 24 if r != row+1 else 36

# 紧急待办
ws1["B18"] = "⚠️ 异常工地预警"
ws1["B18"].font = Font(name="苹方", size=14, bold=True, color="1A1A1A")
ws1.merge_cells("B18:G18")

todo_headers = ["工地编号", "客户", "状态", "约定交期", "超期天数", "负责人"]
for j, h in enumerate(todo_headers):
    c = ws1.cell(row=19, column=2+j, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all

todos = CONFIG["紧急待办"]
for i, t in enumerate(todos):
    r = 20 + i
    for j, v in enumerate(t):
        c = ws1.cell(row=r, column=2+j, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
    if t[4] > 3:
        ws1.cell(row=r, column=6).fill = fill_danger
        ws1.cell(row=r, column=6).font = Font(name="苹方", size=11, bold=True, color=DANGER)

# 状态分布
ws1["I18"] = "状态分布"
ws1["I18"].font = Font(name="苹方", size=14, bold=True, color="1A1A1A")
ws1["I18"].alignment = center
ws1.merge_cells("I18:J18")

for i, (s, n, color) in enumerate(CONFIG["状态分布"]):
    r = 19 + i
    c1 = ws1.cell(row=r, column=9, value=s)
    c1.font = Font(name="苹方", size=11, color=color, bold=True)
    c1.alignment = center
    c2 = ws1.cell(row=r, column=10, value=n)
    c2.font = Font(name="苹方", size=11, bold=True)
    c2.alignment = center
    if i % 2 == 1:
        c1.fill = fill_zebra
        c2.fill = fill_zebra

ws1.freeze_panes = "A4"


# ============================================================
# Sheet 2: 数据源管理
# ============================================================
ws2 = wb.create_sheet("🔌 数据源管理")
ws2.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["数据源表头"], CONFIG["数据源列宽"])):
    c = ws2.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws2.column_dimensions[get_column_letter(j+1)].width = w
ws2.row_dimensions[1].height = 32

for i, src in enumerate(CONFIG["数据源列表"]):
    r = i + 2
    for j, v in enumerate(src):
        c = ws2.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j != 7 else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
    # 状态着色
    if "✅" in src[6]:
        ws2.cell(row=r, column=7).fill = fill_success
        ws2.cell(row=r, column=7).font = Font(name="苹方", size=11, color=SUCCESS, bold=True)
    elif "⚠️" in src[6]:
        ws2.cell(row=r, column=7).fill = fill_warning
        ws2.cell(row=r, column=7).font = Font(name="苹方", size=11, color=WARNING, bold=True)
    elif "❌" in src[6]:
        ws2.cell(row=r, column=7).fill = fill_danger
        ws2.cell(row=r, column=7).font = Font(name="苹方", size=11, color=DANGER, bold=True)

ws2.freeze_panes = "B2"
ws2.auto_filter.ref = f"A1:H{len(CONFIG['数据源列表'])+1}"


# ============================================================
# Sheet 3: 维度配置
# ============================================================
ws3 = wb.create_sheet("⚙️ 维度配置")
ws3.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["维度表头"], CONFIG["维度列宽"])):
    c = ws3.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws3.column_dimensions[get_column_letter(j+1)].width = w
ws3.row_dimensions[1].height = 32

for i, dim in enumerate(CONFIG["维度列表"]):
    r = i + 2
    for j, v in enumerate(dim):
        c = ws3.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j not in [2, 3] else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
    if dim[4] == "✅":
        ws3.cell(row=r, column=5).fill = fill_success
        ws3.cell(row=r, column=5).font = Font(name="苹方", size=11, color=SUCCESS, bold=True)

ws3.freeze_panes = "B2"
ws3.auto_filter.ref = f"A1:F{len(CONFIG['维度列表'])+1}"


# ============================================================
# Sheet 4: 工地管理
# ============================================================
ws4 = wb.create_sheet("🏗️ 工地管理")
ws4.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["工地表头"], CONFIG["工地列宽"])):
    c = ws4.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws4.column_dimensions[get_column_letter(j+1)].width = w
ws4.row_dimensions[1].height = 32

random.seed(7)
for i in range(CONFIG["工地条数"]):
    r = i + 2
    site_no = f"{CONFIG['工地号前缀']}{i+1:03d}"
    area = random.choice(CONFIG["工地区域列表"])
    house = random.choice(CONFIG["工地户型列表"])
    size = random.randint(60, 320)
    phase = random.choice(CONFIG["工地阶段枚举"].split(","))
    progress = random.randint(20, 95) if phase not in ["已竣工", "待开工"] else (100 if phase == "已竣工" else 0)
    foreman = random.choice(CONFIG["工长列表"])
    designer = random.choice(CONFIG["设计师列表"])
    start = (datetime(2026, 1, 1) + timedelta(days=random.randint(0, 130))).strftime("%Y-%m-%d")
    finish = (datetime(2026, 3, 1) + timedelta(days=random.randint(60, 200))).strftime("%Y-%m-%d")
    # 状态逻辑：进度低于 80 且阶段在施工中 → 延期
    if phase in ["拆除", "水电", "泥木", "油漆", "软装"] and progress < 50:
        status = "⚠️ 延期"
    elif phase in ["拆除", "水电", "泥木", "油漆", "软装"] and progress > 90:
        status = "🔧 整改"
    else:
        status = "✅ 正常"
    flag = "无" if "✅" in status else ("工期" if "延期" in status else "验收")

    values = [site_no, f"业主{1000+i}", area, house, size, phase, f"{progress}%",
              foreman, designer, start, finish, status, flag]
    for j, v in enumerate(values):
        c = ws4.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j in [9, 10]:
            c.number_format = 'yyyy-mm-dd'
    if "⚠️" in status:
        ws4.cell(row=r, column=12).fill = fill_warning
        ws4.cell(row=r, column=12).font = Font(name="苹方", size=11, color=WARNING, bold=True)
    elif "❌" in status:
        ws4.cell(row=r, column=12).fill = fill_danger
        ws4.cell(row=r, column=12).font = Font(name="苹方", size=11, color=DANGER, bold=True)

ws4.freeze_panes = "B2"
ws4.auto_filter.ref = f"A1:M{CONFIG['工地条数']+1}"


# ============================================================
# Sheet 5: 客户档案
# ============================================================
ws5 = wb.create_sheet("👥 客户档案")
ws5.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["客户档案表头"], CONFIG["客户档案列宽"])):
    c = ws5.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws5.column_dimensions[get_column_letter(j+1)].width = w
ws5.row_dimensions[1].height = 32

for i, c_data in enumerate(CONFIG["客户档案"]):
    r = i + 2
    for j, v in enumerate(c_data):
        c = ws5.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j not in [8] else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
        if j == 6:
            c.number_format = '¥ #,##0'

ws5.conditional_formatting.add(
    f"H2:H{len(CONFIG['客户档案'])+1}",
    CellIsRule(operator="equal", formula=['"A"'], fill=fill_success,
               font=Font(name="苹方", size=11, color=SUCCESS, bold=True))
)
ws5.conditional_formatting.add(
    f"H2:H{len(CONFIG['客户档案'])+1}",
    CellIsRule(operator="equal", formula=['"B"'], fill=fill_warning,
               font=Font(name="苹方", size=11, color=WARNING, bold=True))
)
ws5.freeze_panes = "B2"


# ============================================================
# Sheet 6: 异常预警
# ============================================================
ws6 = wb.create_sheet("🚨 异常预警")
ws6.sheet_view.showGridLines = False

for j, (h, w) in enumerate(zip(CONFIG["异常预警表头"], CONFIG["异常预警列宽"])):
    c = ws6.cell(row=1, column=j+1, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws6.column_dimensions[get_column_letter(j+1)].width = w
ws6.row_dimensions[1].height = 32

for i, w_data in enumerate(CONFIG["异常预警"]):
    r = i + 2
    for j, v in enumerate(w_data):
        c = ws6.cell(row=r, column=j+1, value=v)
        c.font = font_body
        c.alignment = center if j != 5 else left_align
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra
    # 超期天数 > 3 标红
    days = w_data[3]
    days_num = int(days.split()[0]) if days.split()[0].isdigit() else 0
    if days_num > 3:
        ws6.cell(row=r, column=4).fill = fill_danger
        ws6.cell(row=r, column=4).font = Font(name="苹方", size=11, color=DANGER, bold=True)
    elif days_num > 0:
        ws6.cell(row=r, column=4).fill = fill_warning
        ws6.cell(row=r, column=4).font = Font(name="苹方", size=11, color=WARNING, bold=True)

# 汇总行
r = len(CONFIG["异常预警"]) + 3
ws6.cell(row=r, column=1, value="合计")
ws6.cell(row=r, column=1).fill = fill_header
ws6.cell(row=r, column=1).font = font_header
ws6.cell(row=r, column=1).alignment = center
ws6.merge_cells(start_row=r, end_row=r, start_column=1, end_column=3)
c = ws6.cell(row=r, column=4, value=f"共 {len(CONFIG['异常预警'])} 条待处理")
c.font = Font(name="苹方", size=12, bold=True, color=DANGER)
c.fill = fill_warning
c.alignment = center
for col in range(5, 7):
    ws6.cell(row=r, column=col).fill = fill_header

ws6.freeze_panes = "B2"
ws6.auto_filter.ref = f"A1:F{len(CONFIG['异常预警'])+1}"


# ============================================================
# Sheet 7: 趋势分析
# ============================================================
ws7 = wb.create_sheet("📈 趋势分析")
ws7.sheet_view.showGridLines = False

ws7["B2"] = f"{datetime.now().year} 年家装前端数据趋势"
ws7["B2"].font = Font(name="苹方", size=16, bold=True)
ws7.merge_cells("B2:F2")

monthly_headers = ["月份", "线索数", "营收(元)", "签约数", "流失数", "环比增长"]
for j, h in enumerate(monthly_headers):
    c = ws7.cell(row=4, column=2+j, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border_all
    ws7.column_dimensions[get_column_letter(j+2)].width = 16
ws7.row_dimensions[4].height = 32

monthly = CONFIG["月度数据"]
for i, m in enumerate(monthly):
    r = 5 + i
    ws7.cell(row=r, column=2, value=m[0])
    ws7.cell(row=r, column=3, value=m[1])
    ws7.cell(row=r, column=4, value=m[2]).number_format = '¥ #,##0'
    ws7.cell(row=r, column=5, value=m[3])
    ws7.cell(row=r, column=6, value=m[4])
    # 直接读取环比（已计算）
    ws7.cell(row=r, column=7, value=f"{m[5]}%" if i > 0 else "—")
    for col in range(2, 8):
        c = ws7.cell(row=r, column=col)
        c.font = font_body
        c.alignment = center
        c.border = border_all
        if i % 2 == 1:
            c.fill = fill_zebra

# 同比汇总
r = 5 + len(monthly) + 1
ws7.cell(row=r, column=2, value="本年累计").font = Font(name="苹方", size=12, bold=True, color="FFFFFF")
ws7.cell(row=r, column=2).fill = fill_header
ws7.cell(row=r, column=2).alignment = center
ws7.cell(row=r, column=2).border = border_all
for j, col in enumerate([3, 4, 5, 6]):
    c = ws7.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{4+len(monthly)})")
    c.font = Font(name="苹方", size=12, bold=True)
    c.alignment = center
    c.border = border_all
    c.fill = fill_warning
    if col == 4:
        c.number_format = '¥ #,##0'
ws7.cell(row=r, column=7, value="—").alignment = center
ws7.cell(row=r, column=7).fill = fill_header
ws7.cell(row=r, column=7).font = Font(name="苹方", size=12, bold=True, color="FFFFFF")
ws7.cell(row=r, column=7).border = border_all

ws7.freeze_panes = "B5"

wb.save(OUTPUT)
print(f"✅ 已生成: {OUTPUT}")
print(f"   包含 7 个 Sheet: 数据看板 / 数据源管理 / 维度配置 / 工地管理 / 客户档案 / 异常预警 / 趋势分析")
print(f"   场景: {CONFIG['场景']} · 客户: {CONFIG['客户名']} · 服务商: {CONFIG['服务商名']}")
